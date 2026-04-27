"""
pdf_filler.py — Extract Mix Producer Name, Plant 1 Name, Mix Type,
                 Binder Supplier, Selected Virgin Binder Grade,
                 JMF Number, Calibration Number, aggregate bin %/Gsb,
                 RAP pile/%, and AC%/Binder Gb
                 from a Marshall Mix Design PDF.

Usage:
    python pdf_filler.py <pdf_path> <output_json_path>
"""

import sys
import re
import json


def extract_fields(pdf_path: str) -> dict:
    try:
        import pdfplumber
    except ImportError:
        return {"producer": "", "plant1": "", "mix_type": "",
                "binder_supplier": "", "binder_grade": "",
                "jmf_number": "", "calib_number": "",
                "error": "pdfplumber not installed"}

    producer = ""
    plant1 = ""
    mix_type = ""
    binder_supplier = ""
    binder_grade = ""
    jmf_number = ""
    calib_number = ""
    virgin_binder = ""
    rap_binder = ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            page1_text = pdf.pages[0].extract_text() or ""
            page2_text = pdf.pages[1].extract_text() if len(pdf.pages) > 1 else ""
            # Collect all page text for JMF / Calib search
            all_pages_text = "\n".join(
                (p.extract_text() or "") for p in pdf.pages
            )

        # --- Page 1 fields ---

        # "Mix Producer Name: <value>" — ends at newline
        m = re.search(r"Mix Producer Name:\s*(.+)", page1_text)
        if m:
            producer = m.group(1).strip()

        # "Plant 1 Name: <value>" — ends at newline
        m = re.search(r"Plant 1 Name:\s*(.+)", page1_text)
        if m:
            plant1 = m.group(1).strip()

        # "Mix Type <value>" — appears on the line starting with "Mix info:"
        # Format: "Mix info: Mix Type 441 Type 2 Intermediate"
        # Capture everything after "Mix Type" until end of line
        m = re.search(r"Mix Type\s+(.+?)(?:\s{2,}|\n|$)", page1_text)
        if m:
            mix_type = m.group(1).strip()

        # --- Page 2 fields ---

        # "Selected Virgin Binder Grade PG 64-22 Neat Producer Name Cargill"
        # Capture the grade up to "Neat" or "Producer Name" (whichever comes first)
        m = re.search(r"Selected Virgin Binder Grade\s+(.+?)\s+(?:Neat\b|Producer Name)", page2_text)
        if m:
            binder_grade = m.group(1).strip()
        else:
            # Fallback: capture until end of line
            m = re.search(r"Selected Virgin Binder Grade\s+(.+)", page2_text)
            if m:
                binder_grade = m.group(1).strip()

        # "Binder Supplier S&S Terminal - Rayland, OH Brand Name Anova"
        # Capture supplier up to "Brand Name"
        m = re.search(r"Binder Supplier\s+(.+?)\s+Brand Name", page2_text)
        if m:
            binder_supplier = m.group(1).strip()
        else:
            # Fallback: capture until end of line
            m = re.search(r"Binder Supplier\s+(.+)", page2_text)
            if m:
                binder_supplier = m.group(1).strip()

        # --- % Virgin Binder and % Binder from RAP (page 1) ---

        m = re.search(r'%\s*Virgin\s*Binder\s+([\d.]+)', page2_text, re.IGNORECASE)
        if m:
            virgin_binder = m.group(1).strip()
        else:
            virgin_binder = ""

        m = re.search(r'%\s*Binder\s+from\s+RAP\s+([\d.]+)', page2_text, re.IGNORECASE)
        if m:
            rap_binder = m.group(1).strip()
        else:
            rap_binder = ""

        # --- JMF Number and Calibration Number (search all pages) ---

        # JMF number: B-number immediately before the "~" separator
        # e.g. "B260208 ~" or "B260208~"
        m = re.search(r"(B\d+)\s*~", all_pages_text)
        if m:
            jmf_number = m.group(1).strip()

        # Calibration number: digits following "Calib#"
        # e.g. "Calib# 60208" or "Calib#60208"
        m = re.search(r"Calib#\s*(\d+)", all_pages_text)
        if m:
            calib_number = m.group(1).strip()

    except Exception as e:
        return {"producer": "", "plant1": "", "mix_type": "",
                "binder_supplier": "", "binder_grade": "",
                "jmf_number": "", "calib_number": "", "error": str(e)}

    return {
        "producer": producer,
        "plant1": plant1,
        "mix_type": mix_type,
        "binder_supplier": binder_supplier,
        "binder_grade": binder_grade,
        "jmf_number": jmf_number,
        "calib_number": calib_number,
        "virgin_binder": virgin_binder,
        "rap_binder": rap_binder,
    }


def abbreviate_material(name: str) -> str:
    """Return a short label (max 4 chars) for a material name.
    Examples: 'Natural Gravel 057' -> 'G57', 'Natural Sand SD5' -> 'NS',
              'Limestone Sand' -> 'LSS', 'Baghouse Fines' -> 'BHF'.
    """
    n = name.strip()
    if not n:
        return ""

    # Natural Gravel / Gravel NNNN  →  G + up to 3 digits (strip leading zeros)
    m = re.match(r'(?:Natural\s+)?Gravel\s+0*(\d+)', n, re.IGNORECASE)
    if m:
        return ("G" + m.group(1))[:4]

    # Limestone Sand  →  LSS
    if re.search(r'limestone\s+sand', n, re.IGNORECASE):
        return "LSS"

    # Natural Sand (any suffix)  →  always NS
    if re.search(r'natural\s+sand', n, re.IGNORECASE):
        return "NS"

    # Baghouse Fines  →  BHF
    if re.search(r'baghouse', n, re.IGNORECASE):
        return "BHF"

    # Crushed Limestone  →  CLS
    if re.search(r'crushed\s+limestone', n, re.IGNORECASE):
        return "CLS"

    # Limestone (alone)  →  LS
    if re.search(r'limestone', n, re.IGNORECASE):
        return "LS"

    # Stone Sand  →  StS
    if re.search(r'stone\s+sand', n, re.IGNORECASE):
        return "StS"

    # Screenings  →  SCR
    if re.search(r'screening', n, re.IGNORECASE):
        return "SCR"

    # Crushed Stone  →  CS
    if re.search(r'crushed\s+stone', n, re.IGNORECASE):
        return "CS"

    # Manufactured Sand  →  MfS
    if re.search(r'manufactured\s+sand', n, re.IGNORECASE):
        return "MfS"

    # Slag  →  SLG
    if re.search(r'slag', n, re.IGNORECASE):
        return "SLG"

    # Fallback: initials of each word, capped at 4 chars
    words = n.split()
    return "".join(w[0].upper() for w in words if w)[:4]


def parse_page3_materials(pdf_path):
    import pdfplumber, re
    from openpyxl.utils import get_column_letter

    with pdfplumber.open(pdf_path) as pdf:
        page3_text = pdf.pages[2].extract_text() or ""
        page2_text = pdf.pages[1].extract_text() or ""

    lines = page3_text.split('\n')

    def find_section(name):
        for i, l in enumerate(lines):
            if name in l:
                return i
        return None

    coarse_i = find_section('Coarse Aggregates')
    fine_i   = find_section('Fine Aggregates')
    bag_i    = find_section('Baghouse Fines')
    rap_i    = find_section('RAP')
    blend_i  = find_section('Blend Gsb')

    def parse_agg_line(line):
        tokens = line.strip().split()
        if len(tokens) < 8:
            return None
        try:
            pct = float(tokens[0])
            gsb = float(tokens[-1])
            float(tokens[-2])
        except ValueError:
            return None
        if pct == 0.0:
            return None
        size      = tokens[-3]
        type_name = tokens[-5] + ' ' + tokens[-4]
        producer  = ' '.join(tokens[2:-5])
        return {"material": type_name + ' ' + size, "producer": producer, "pct": pct, "gsb": gsb}

    def parse_bag_line(line):
        tokens = line.strip().split()
        if len(tokens) < 5:
            return None
        try:
            pct = float(tokens[0])
            gsb = float(tokens[-1])
        except ValueError:
            return None
        if pct == 0.0:
            return None
        type_name = tokens[2] + ' ' + tokens[3]
        producer  = ' '.join(tokens[4:-1])
        return {"material": type_name, "producer": producer, "pct": pct, "gsb": gsb}

    def parse_rap_line(line):
        match = re.search(r'Method \d+\s+(.+?)\s+[A-Z]+/[A-Z]+\s+(\d+\.\d+)', line)
        if match:
            tokens = line.strip().split()
            try: pct = float(tokens[0])
            except: pct = 0.0
            return {
                "pile": match.group(1).strip().replace('"', 'in.'),
                "pct":  pct,
                "gse":  float(match.group(2))
            }
        return None

    # Parse each section
    coarse_items = []
    if coarse_i is not None and fine_i is not None:
        for line in lines[coarse_i+1:fine_i]:
            r = parse_agg_line(line)
            if r:
                r["item"] = "703.50"
                coarse_items.append(r)

    fine_items = []
    if fine_i is not None and bag_i is not None:
        for line in lines[fine_i+1:bag_i]:
            r = parse_agg_line(line)
            if r:
                r["item"] = "703.05"
                fine_items.append(r)

    bag_items = []
    if bag_i is not None and rap_i is not None:
        for line in lines[bag_i+1:rap_i]:
            r = parse_bag_line(line)
            if r:
                r["item"] = "703.05"
                bag_items.append(r)

    rap_data = {"pile": "", "pct": 0.0, "gse": 0.0}
    if rap_i is not None:
        end = blend_i if blend_i else len(lines)
        for line in lines[rap_i+1:end]:
            r = parse_rap_line(line)
            if r:
                rap_data = r
                break

    # Compute dynamic bin column assignments
    # Coarse fills F,G,H,I (base col 6), skip 1, Fine fills next, Baghouse right after Fine
    BASE_COL      = 6  # column F
    num_coarse    = len(coarse_items)
    fine_start    = BASE_COL + num_coarse + 1   # +1 for the skip
    bag_start     = fine_start + len(fine_items)

    aggs = []
    for i, item in enumerate(coarse_items):
        aggs.append({**item, "col": get_column_letter(BASE_COL + i)})
    for i, item in enumerate(fine_items):
        aggs.append({**item, "col": get_column_letter(fine_start + i)})
    for i, item in enumerate(bag_items):
        aggs.append({**item, "col": get_column_letter(bag_start + i)})

    # Build 6-slot material list (same as before, now includes pct/gsb)
    empty = {"material": "", "producer": "", "item": "", "pct": "", "gsb": ""}
    slots = []
    for r in coarse_items[:4]: slots.append(r)
    slots.append(empty)
    for r in fine_items[:4]:   slots.append(r)
    for r in bag_items:        slots.append(r)
    while len(slots) < 6:     slots.append(empty)
    slots = slots[:6]

    # Extract AC% and Binder Gb from page 2
    ac_pct    = ""
    binder_gb = ""
    ac_match  = re.search(r'% Binder Content.*?Opt.*?Air Voids\s+([\d.]+)', page2_text, re.IGNORECASE)
    if ac_match:
        ac_pct = ac_match.group(1)
    gb_match  = re.search(r'Binder Gb\s+([\d.]+)', page2_text, re.IGNORECASE)
    if gb_match:
        binder_gb = gb_match.group(1)

    # Build result dict — all values stored as strings for VBA JSON parser compatibility
    result = {
        "rap_pile": rap_data["pile"],
        "rap_pct":  str(rap_data["pct"]),
        "rap_gse":  str(rap_data["gse"]),
        "ac_pct":   ac_pct,
        "binder_gb": binder_gb,
        "agg_count": str(len(aggs))
    }
    for i, s in enumerate(slots, 1):
        result[f"material_{i}"] = s["material"]
        result[f"producer_{i}"] = s["producer"]
        result[f"item_{i}"]     = s["item"]
    for i, a in enumerate(aggs, 1):
        result[f"agg_{i}_col"]    = a["col"]
        result[f"agg_{i}_pct"]    = str(a["pct"])
        result[f"agg_{i}_gsb"]    = str(a["gsb"])
        result[f"agg_{i}_abbrev"] = abbreviate_material(a["material"])

    return result


def parse_odot_spec_band(pdf_path: str) -> dict:
    """Extract ODOT SPEC. BAND sieve table from page 2 of the PDF.

    Returns a dict with keys sieve_N_jmf and sieve_N_mr for N=0..12.
    sieve_N_jmf is the JMF % pass as a string.
    sieve_N_mr  is "Low / High" as a string, or "" if Low/High are absent.

    Sieve index order (matches the 13-row table):
      0=2", 1=1-1/2", 2=1", 3=3/4", 4=1/2", 5=3/8",
      6=#4,  7=#8,   8=#16, 9=#30, 10=#50, 11=#100, 12=#200
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    # Each entry: (index, search_keyword)
    # The keyword is a unique substring that appears in the sieve label line.
    SIEVES = [
        (0,  r'2"\s*\(50'),        # 2" (50.8)
        (1,  r'1-1/2"\s*\(38'),    # 1-1/2" (38.1)
        (2,  r'1"\s*\(25'),        # 1" (25.4)
        (3,  r'3/4"\s*\(19\)'),    # 3/4" (19)
        (4,  r'1/2"\s*\(12'),      # 1/2" (12.7)
        (5,  r'3/8"\s*\(9'),       # 3/8" (9.5)
        (6,  r'#4\s*\(4'),         # #4 (4.75)
        (7,  r'#8\s*\(2'),         # #8 (2.36)
        (8,  r'#16\s*\(1\.1'),     # #16 (1.18)
        (9,  r'#30\s*\(0\.6\)'),   # #30 (0.6)
        (10, r'#50\s*\(0\.3\)'),   # #50 (0.3)
        (11, r'#100\s*\(0\.1'),    # #100 (0.15)
        (12, r'#200\s*\(0\.0'),    # #200 (0.075)
    ]

    result = {}
    for idx in range(13):
        result[f"sieve_{idx}_jmf"] = ""
        result[f"sieve_{idx}_mr"]  = ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) < 2:
                return result
            page2_text = pdf.pages[1].extract_text() or ""

        # Split into lines for per-line matching
        lines = page2_text.split('\n')

        for (idx, pattern) in SIEVES:
            for line in lines:
                if re.search(pattern, line):
                    # The line format: '<label> (<mm>)  <jmf> [<low> <high>] [text...]'
                    # Strip everything up to and including the closing ')' of the mm value,
                    # then extract numbers from what remains.
                    # e.g. '2" (50.8) 100 100 100 Fibers Brand Name'
                    #   -> after ')': ' 100 100 100 Fibers Brand Name'
                    m_paren = re.search(r'\)', line)
                    if m_paren:
                        after_paren = line[m_paren.end():]
                        trailing_nums = re.findall(r'\d+\.?\d*', after_paren)
                        if trailing_nums:
                            result[f"sieve_{idx}_jmf"] = trailing_nums[0]
                            if len(trailing_nums) >= 3:
                                result[f"sieve_{idx}_mr"] = f"{trailing_nums[1]} / {trailing_nums[2]}"
                            else:
                                result[f"sieve_{idx}_mr"] = ""
                    break  # found the line for this sieve, move to next

    except Exception as e:
        result["sieve_error"] = str(e)

    return result


def main():
    if len(sys.argv) != 3:
        print("Usage: python pdf_filler.py <pdf_path> <output_json_path>")
        sys.exit(1)

    pdf_path = sys.argv[1]
    output_path = sys.argv[2]

    result = extract_fields(pdf_path)

    try:
        materials = parse_page3_materials(pdf_path)
        result.update(materials)
    except Exception as e:
        result["materials_error"] = str(e)

    try:
        sieve_data = parse_odot_spec_band(pdf_path)
        result.update(sieve_data)
    except Exception as e:
        result["sieve_error"] = str(e)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
